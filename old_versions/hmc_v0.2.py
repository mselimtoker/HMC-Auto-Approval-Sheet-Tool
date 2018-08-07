import pyexcel
from openpyxl import Workbook, load_workbook
import os
from openpyxl.styles import Border, Side

apv_position = 'F22'

a2_reference_table = 'Reference table A2 Engine.xlsx'

u2_reference_table = 'Reference table U2 Engine.xlsx'

sif_folder_name = 'Sample_SIF_Files'
sif_tab_name = 'Approvals'

is_multical = False

template_name = 'temp.xlsx'
multical_template_name = 'multi_temp.xlsx'

temp_start_row = 2


def get_from_sif(dds_serial):
    file_path = sif_folder_name + '/' + dds_serial + '.xls'
    try:
        s = pyexcel.get_sheet(file_name=file_path, sheet_name=sif_tab_name)
    except:
        print dds_serial + '.xls not found.'
    return s.row[21][5]


def get_and_set_xl(row, i, type):
    if type == 'U2' or 'u2':
       #if (sheet.row[row][4] or sheet.row[row][3] or sheet.row[row][10] or sheet.row[row][11] or sheet.row[row][0] or
       #    sheet.row[row][6] or sheet.row[row][5] or sheet.row[row][16] or sheet.row[row][17] or
       #    sheet.row[row][15]) == "" or " ":

       #    return 1
       #else:
        ws['D' + str(i)] = sheet.row[row][4]  # emission_level
        ws['E' + str(i)] = sheet.row[row][3][0:2]  # veh
        ws['F' + str(i)] = sheet.row[row][10]  # market
        ws['G' + str(i)] = sheet.row[row][11]  # gearbox
        ws['I' + str(i)] = sheet.row[row][0]  # hmc_ecu_part_no
        ws['L' + str(i)] = sheet.row[row][6]  # hmc_rom_id
        ws['M' + str(i)] = sheet.row[row][5]  # hmc_cal_id
        ws['O' + str(i)] = sheet.row[row][16]  # cvn_number
        ws['Q' + str(i)] = sheet.row[row][17]  # dds_rom_id
        ws['R' + str(i)] = sheet.row[row][15]  # ulp_file
        ws['N' + str(i)] = get_from_sif(sheet.row[row][15])  # app_checksum

    elif type == 'A2' or 'a2':
      ## if (sheet.row[row][4] or sheet.row[row][3] or sheet.row[row][13] or sheet.row[row][14] or sheet.row[row][0] or
      #     sheet.row[row][8] or sheet.row[row][7] or sheet.row[row][18] or sheet.row[row][19] or
      #     sheet.row[row][17]) == "" or " ":

      #     return 1
      #  else:
        ws['D' + str(i)] = sheet.row[row][4]  # emission_level
        ws['E' + str(i)] = sheet.row[row][3][0:2]  # veh
        ws['F' + str(i)] = sheet.row[row][13]  # market
        ws['G' + str(i)] = sheet.row[row][14]  # gearbox
        ws['I' + str(i)] = sheet.row[row][0]  # hmc_ecu_part_no
        ws['L' + str(i)] = sheet.row[row][8]  # hmc_rom_id
        ws['M' + str(i)] = sheet.row[row][7]  # hmc_cal_id
        ws['O' + str(i)] = sheet.row[row][18]  # cvn_number
        ws['Q' + str(i)] = sheet.row[row][19]  # dds_rom_id
        ws['R' + str(i)] = sheet.row[row][17]  # ulp_file
        ws['N' + str(i)] = get_from_sif(sheet.row[row][15])  # app_checksum




while True:
    print "Select your process type. A2/U2: "
    process_type = raw_input()

    if process_type == 'A2' or process_type == 'a2':
        selected_xl = a2_reference_table
        selected_template = template_name
        break

    elif process_type == 'U2' or process_type == 'u2':
        selected_xl = u2_reference_table

        while True:
            print "Is it multi calibration process? Y/N: "
            multical_input = raw_input()

            if multical_input == 'Y' or multical_input == 'y':
                is_multical = True
                selected_template = multical_template_name
                break

            elif multical_input == 'N' or multical_input == 'n':
                is_multical = False
                selected_template = template_name
                break

            else:
                print "Invalid entry!! Please type Y or N."
        break
    else:
        print "Invalid process type !! Please enter A2 or U2."

if os.path.isfile(selected_xl):
    if os.path.isfile(selected_template):

        book = pyexcel.get_book(file_name=selected_xl)
        is_found = False

        while True:
            print "Please enter tab name for " + selected_xl
            tab_input = raw_input()
            for sheets in book:
                if tab_input == sheets.name:
                    active_tab_name = tab_input
                    is_found = True

            if is_found:
                break
            else:
                print tab_input + " not found. Please check tab name."

        while True:
            try:
                start_row = int(input("Enter start position of selected row. (Include): "))
                break
            except:
                print "You must enter a number"

        while True:
            try:
                end_row = int(input("Enter end position of selected row. (Include): "))
                break
            except:
                print "You must enter a number"

        sheet = pyexcel.get_sheet(file_name=selected_xl, sheet_name=active_tab_name)

        wb = load_workbook(filename=selected_template)
        ws = wb.worksheets[0]

        i = temp_start_row

        for rows in range(start_row - 1, end_row):
            get_and_set_xl(rows, i, process_type)
            i += 1

        for row in range(1, 32):
            for col in range(1, 19):
                ws.cell(row=row, column=col).border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                                             top=Side(style='thin'), bottom=Side(style='thin'))

        wb.save('output.xlsx')
    else:
        print selected_template + " not found."
else:
    print selected_xl + " not found."
