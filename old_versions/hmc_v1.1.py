import pyexcel
from openpyxl import Workbook, load_workbook
import os
from openpyxl.styles import Border, Side

apv_position = 'F22'

a2_reference_table = 'Reference table A2 Engine.xlsx'

u2_reference_table = 'Reference table U2 Engine.xlsx'

sif_folder_name = 'Sample_SIF_Files'
sif_tab_name = 'Approvals'

is_multical = False

template_name = 'temp.xlsx'
multical_template_name = 'multi_temp.xlsx'

temp_start_row = 2

def get_from_sif(dds_serial):
    file_path = sif_folder_name + '/' + dds_serial + '.xls'
    if os.path.isfile(file_path):
        s = pyexcel.get_sheet(file_name=file_path, sheet_name=sif_tab_name)
        if s.row[21][5] == "" or s.row[21][5] == " ":
            print "Code not found in " + dds_serial + ".xls"
            return False
        else:
            return s.row[21][5]
    else:
        print dds_serial + '.xls not found.'
        return False




    return s.row[21][5]

def check_missing_value(*args):
    for arg in args:
        if arg == "" or arg ==" ":
            return True
            break
        else:
            return False

def get_and_set_xl(row, i, type):
    sr = sheet.row[row]

    if type == 'A2' or type == 'a2':
        if check_missing_value(sr[4], sr[3], sr[10], sr[11], sr[0], sr[6], sr[5], sr[16], sr[17], sr[15]):
            print selected_xl + " has missing values."
            return False
        else:
            g = get_from_sif(sr[15])
            if g == False:
                return False
            else:
                ws['N' + str(i)] = g  # app_checksum
                ws['D' + str(i)] = sr[4]  # emission_level
                ws['E' + str(i)] = sr[3][0:2]  # veh
                ws['F' + str(i)] = sr[10]  # market
                ws['G' + str(i)] = sr[11]  # gearbox
                ws['I' + str(i)] = sr[0]  # hmc_ecu_part_no
                ws['L' + str(i)] = sr[6]  # hmc_rom_id
                ws['M' + str(i)] = sr[5]  # hmc_cal_id
                ws['O' + str(i)] = sr[16]  # cvn_number
                ws['Q' + str(i)] = sr[17]  # dds_rom_id
                ws['R' + str(i)] = sr[15]  # ulp_file
                return True




    elif type == 'U2' or type == 'u2':
        if check_missing_value(sr[4], sr[3], sr[13], sr[14], sr[0], sr[8], sr[7], sr[18], sr[19], sr[17]):
            print selected_xl + " has missing values."
            return False
        else:

            g = get_from_sif(sr[17])
            if g is False:
                return False
            else:
                ws['N' + str(i)] = g  # app_checksum
                ws['D' + str(i)] = sr[4]  # emission_level
                ws['E' + str(i)] = sr[3][0:2]  # veh
                ws['F' + str(i)] = sr[13]  # market
                ws['G' + str(i)] = sr[14]  # gearbox
                ws['I' + str(i)] = sr[0]  # hmc_ecu_part_no
                ws['L' + str(i)] = sr[8]  # hmc_rom_id
                ws['M' + str(i)] = sr[7]  # hmc_cal_id
                ws['O' + str(i)] = sr[18]  # cvn_number
                ws['Q' + str(i)] = sr[19]  # dds_rom_id
                ws['R' + str(i)] = sr[17]  # ulp_file
                return True

i = temp_start_row

while True:
    print "Select your process type. A2/U2: "
    process_type = raw_input()

    if process_type == 'A2' or process_type == 'a2':
        selected_xl = a2_reference_table
        selected_template = template_name
        break

    elif process_type == 'U2' or process_type == 'u2':
        selected_xl = u2_reference_table

        while True:
            print "Is it multi calibration process? Y/N: "
            multical_input = raw_input()

            if multical_input == 'Y' or multical_input == 'y':
                is_multical = True
                selected_template = multical_template_name
                break

            elif multical_input == 'N' or multical_input == 'n':
                is_multical = False
                selected_template = template_name
                break

            else:
                print "Invalid entry!! Please type Y or N."
        break
    else:
        print "Invalid process type !! Please enter A2 or U2."

wb = load_workbook(filename=selected_template)
ws = wb.worksheets[0]

while True:
    if os.path.isfile(selected_xl):
        if os.path.isfile(selected_template):

            book = pyexcel.get_book(file_name=selected_xl)
            is_found = False

            while True:
                print "Please enter tab name for " + selected_xl
                tab_input = raw_input()
                for sheets in book:
                    if tab_input == sheets.name:
                        active_tab_name = tab_input
                        is_found = True

                if is_found:
                    break
                else:
                    print tab_input + " not found. Please check tab name."

            while True:
                try:
                    start_row = int(input("Enter start position of selected row. (Include): "))
                    break
                except:
                    print "You must enter a number"

            while True:
                try:
                    end_row = int(input("Enter end position of selected row. (Include): "))
                    break
                except:
                    print "You must enter a number"

            sheet = pyexcel.get_sheet(file_name=selected_xl, sheet_name=active_tab_name)


            for rows in range(start_row - 1, end_row):
                val = get_and_set_xl(rows, i, process_type)
                if val is False:
                    break
                i += 1
            for row in range(1, 200):
                for col in range(1, 19):
                    ws.cell(row=row, column=col).border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                                                   top=Side(style='thin'), bottom=Side(style='thin'))
        else:
            print selected_template + " not found."
    else:
        print selected_xl + " not found."
    print "Values added to excel."
    while True:
        print "Do you want to repeat process? Y/N"
        repeat = raw_input()

        if repeat == "N" or repeat == "n":
            break
        elif repeat == "Y" or repeat == "y":
            print "Process starts again"
            break
        else:
            print "Invalid entry! Please enter Y or N."

    if repeat == "N" or repeat == "n":
        if not val == False:
            wb.save('outputt.xlsx')
            print "Process completed."
        break
